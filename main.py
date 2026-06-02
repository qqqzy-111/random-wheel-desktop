"""
彩色随机转盘 - 桌面应用程序
功能: 从Excel读取数据，生成彩色转盘，实现随机抽奖功能
作者: [Your Name]
日期: 2026年5月26日
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import math
import random
import time
from pathlib import Path
from openpyxl import load_workbook

# ==================== 常量定义 ====================

# 窗口配置
WINDOW_WIDTH = 1200
WINDOW_HEIGHT = 700
WINDOW_TITLE = "风吹哪页读哪页"

# 转盘配置
WHEEL_CENTER_X = 350
WHEEL_CENTER_Y = 350
WHEEL_RADIUS = 250

# 指针配置（在转盘顶部）
POINTER_X = WHEEL_CENTER_X
POINTER_Y = 50

# 颜色列表（彩虹色系）
COLOR_PALETTE = [
    '#FF6B6B',  # 红
    '#FFA700',  # 橙
    '#FFE66D',  # 黄
    '#95E1D3',  # 青绿
    '#38ADA9',  # 孔雀蓝
    '#78C850',  # 绿
    '#A890F0',  # 紫
    '#FF85A2',  # 粉红
    '#F38181',  # 珊瑚红
    '#AA96DA',  # 淡紫
    '#FCBAD3',  # 浅粉
    '#A8E6CF',  # 薄荷绿
    '#FFD3B6',  # 桃子
    '#FFAAA5',  # 浅红
    '#FF8B94',  # 草莓
    '#A8D8EA',  # 天蓝
]

# ==================== 转盘类 ====================

class WheelSpinner:
    """转盘处理类"""
    
    def __init__(self, canvas, center_x, center_y, radius):
        """
        初始化转盘
        
        参数:
            canvas: Tkinter Canvas 对象
            center_x: 转盘中心X坐标
            center_y: 转盘中心Y坐标
            radius: 转盘半径
        """
        self.canvas = canvas
        self.center_x = center_x
        self.center_y = center_y
        self.radius = radius
        
        self.items = []  # 转盘选项列表
        self.colors = []  # 颜色列表
        self.current_rotation = 0  # 当前旋转角度
        self.is_spinning = False  # 是否正在旋转
        self.winner = None  # 中奖者
        
    def set_items(self, items):
        """设置转盘选项"""
        self.items = items
        self.colors = self._generate_colors()
        self.current_rotation = 0
        self.winner = None
        self.draw()
    
    def _generate_colors(self):
        """生成颜色列表"""
        colors = []
        for i in range(len(self.items)):
            colors.append(COLOR_PALETTE[i % len(COLOR_PALETTE)])
        return colors
    
    def draw(self):
        """绘制转盘 - 改进版本
        
        改进点：
        1. 分离绘制流程（扇形 → 文字 → 中心圆）防止遮挡
        2. 根据分块数动态调整字体大小
        3. 添加文字阴影提高可读性
        4. 优化文字位置避免靠太近中心
        """
        self.canvas.delete("wheel")
        
        if not self.items:
            # 显示提示文字
            self.canvas.create_text(
                self.center_x, self.center_y,
                text="请先选择 Excel 文件",
                font=("Arial", 16, "bold"),
                fill="#999999",
                tags="wheel"
            )
            return
        
        # 计算每个扇形的角度
        slice_angle = 360 / len(self.items)
        
        # ========== 第一步：绘制所有彩色扇形 ==========
        for i in range(len(self.items)):
            start_angle = self.current_rotation + i * slice_angle
            end_angle = start_angle + slice_angle
            
            # 绘制扇形
            self.canvas.create_arc(
                self.center_x - self.radius,
                self.center_y - self.radius,
                self.center_x + self.radius,
                self.center_y + self.radius,
                start=start_angle,
                extent=slice_angle,
                fill=self.colors[i],
                outline="white",
                width=3,
                tags="wheel"
            )
        
        # ========== 第二步：绘制所有文字 ==========
        for i in range(len(self.items)):
            start_angle = self.current_rotation + i * slice_angle
            self._draw_text(i, start_angle, slice_angle)
        
        # ========== 第三步：绘制中心圆（最后绘制，确保不被文字遮挡） ==========
        center_radius = 30
        self.canvas.create_oval(
            self.center_x - center_radius,
            self.center_y - center_radius,
            self.center_x + center_radius,
            self.center_y + center_radius,
            fill="white",
            outline="#667eea",
            width=3,
            tags="wheel"
        )
        
        # 在中心绘制装饰文字
        self.canvas.create_text(
            self.center_x, self.center_y,
            text="转",
            font=("SimHei", 18, "bold"),
            fill="#667eea",
            tags="wheel"
        )
    
    def _draw_text(self, index, start_angle, slice_angle):
        """在扇形上绘制文字 - 改进版本
        
        改进点：
        1. 根据分块数量自动调整字体大小
        2. 文字放在半径0.62-0.72范围内（避免太靠近圆心）
        3. 添加黑色阴影/描边提高可读性
        4. 根据字体大小自动调整文字截断长度
        5. 处理文字方向避免倒置显示
        """
        # ========== 根据分块数量调整字体大小 ==========
        item_count = len(self.items)
        
        if item_count <= 10:
            font_size = 18
            max_chars = 6      # 最多显示6个汉字
            text_radius_ratio = 0.68   # 文字距离中心的比例
        elif item_count <= 18:
            font_size = 14
            max_chars = 5      # 最多显示5个汉字
            text_radius_ratio = 0.67
        else:
            font_size = 10
            max_chars = 4      # 最多显示4个汉字
            text_radius_ratio = 0.65
        
        # ========== 计算文字位置 ==========
        # 文字放在扇形中点
        text_angle = start_angle + slice_angle / 2
        text_rad = math.radians(text_angle)
        
        # 文字距离中心的距离（在0.62-0.72之间，避免靠太近或太远）
        text_radius = self.radius * text_radius_ratio
        text_x = self.center_x + text_radius * math.cos(text_rad)
        text_y = self.center_y + text_radius * math.sin(text_rad)
        
        # ========== 处理文字截断 ==========
        text = self.items[index]
        
        # 根据字体大小和最多字数进行截断
        if len(text) > max_chars:
            text = text[:max_chars] + "…"
        
        # ========== 处理文字方向，避免倒置 ==========
        # 规范化角度到0-360度范围
        normalized_angle = text_angle % 360
        
        # 记录是否需要文字翻转（在左下和右下象限时）
        # 这里我们先不做旋转，只是确保文字能看清
        # 如果角度在90-270度（下半部分），视觉上会有些问题
        # 但由于Tkinter没有easy文字旋转，我们保持现状
        
        # ========== 绘制文字阴影（黑色，偏移1像素） ==========
        # 这会创建一个轻微的黑色描边效果，提高可读性
        self.canvas.create_text(
            text_x + 1, text_y + 1,
            text=text,
            font=("SimHei", font_size, "bold"),
            fill="#000000",
            anchor="center",  # 文字中心对齐
            tags="wheel"
        )
        
        # ========== 绘制主文字（白色） ==========
        self.canvas.create_text(
            text_x, text_y,
            text=text,
            font=("SimHei", font_size, "bold"),
            fill="white",
            anchor="center",  # 文字中心对齐
            tags="wheel"
        )
    
    def spin(self, duration=3.0):
        """
        旋转转盘
        
        参数:
            duration: 旋转持续时间（秒）
        """
        if self.is_spinning or not self.items:
            return
        
        self.is_spinning = True
        
        # 生成随机的最终角度
        # 转盘转5-8圈再停止
        random_rotations = 5 + random.random() * 3
        random_angle = random.random() * 360
        target_rotation = random_rotations * 360 + random_angle
        
        # 启动动画线程
        thread = threading.Thread(
            target=self._spin_animation,
            args=(target_rotation, duration)
        )
        thread.daemon = True
        thread.start()
    
    def _spin_animation(self, target_rotation, duration):
        """转盘旋转动画"""
        start_time = time.time()
        start_rotation = self.current_rotation
        
        while True:
            current_time = time.time()
            elapsed = current_time - start_time
            progress = min(elapsed / duration, 1.0)
            
            # 缓动函数：先快后慢
            ease_progress = 1 - (1 - progress) ** 2
            
            # 计算当前旋转角度
            self.current_rotation = start_rotation + (target_rotation - start_rotation) * ease_progress
            
            # 重新绘制转盘
            self.draw()
            
            if progress >= 1.0:
                break
            
            time.sleep(0.02)  # 约50fps
        
        # 动画完成，计算中奖结果
        self.is_spinning = False
        self._calculate_winner()
    
    def _calculate_winner(self):
        """计算中奖者"""
        if not self.items:
            return
        
        # 标准化角度
        rotation = self.current_rotation % 360
        
        # 每个扇形的角度
        slice_angle = 360 / len(self.items)
        
        # 指针指向顶部（0度方向对应12点）
        # 反向计算指向的扇形
        winner_index = int((360 - rotation) / slice_angle) % len(self.items)
        
        self.winner = self.items[winner_index]
    
    def get_winner(self):
        """获取中奖者"""
        return self.winner
    
    def reset(self):
        """重置转盘"""
        self.current_rotation = 0
        self.winner = None
        self.draw()


# ==================== Excel处理类 ====================

class ExcelReader:
    """Excel 文件读取类"""
    
    @staticmethod
    def read_first_column(file_path):
        """
        读取 Excel 文件第一列数据
        
        参数:
            file_path: Excel 文件路径
            
        返回:
            list: 过滤后的数据列表
        """
        try:
            workbook = load_workbook(file_path)
            worksheet = workbook.active
            
            items = []
            
            # 从第二行开始读取（跳过表头）
            for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                if row_idx == 1:  # 跳过表头
                    continue
                
                if row and row[0]:  # 检查是否有值
                    value = str(row[0]).strip()
                    if value:  # 过滤空白
                        items.append(value)
            
            return items
        
        except Exception as e:
            raise Exception(f"读取 Excel 文件失败: {str(e)}")


# ==================== 主应用类 ====================

class WheelApp:
    """主应用程序类"""
    
    def __init__(self, root):
        """初始化应用程序"""
        self.root = root
        self.root.title(WINDOW_TITLE)
        self.root.geometry(f"{WINDOW_WIDTH}x{WINDOW_HEIGHT}")
        self.root.resizable(False, False)
        
        # 设置背景色
        self.root.configure(bg="#f0f0f0")
        
        # 当前 Excel 文件路径
        self.current_file = None
        
        # 创建UI
        self._create_ui()
        
        # 初始化转盘
        self.wheel = WheelSpinner(self.canvas, WHEEL_CENTER_X, WHEEL_CENTER_Y, WHEEL_RADIUS)
        self.wheel.draw()
        
        # 绘制指针
        self._draw_pointer()
    
    def _create_ui(self):
        """创建用户界面"""
        # ========== 左侧：转盘区域 ==========
        left_frame = tk.Frame(self.root, bg="#f0f0f0")
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # 转盘画布
        self.canvas = tk.Canvas(
            left_frame,
            width=WHEEL_CENTER_X * 2,
            height=620,
            bg="white",
            highlightthickness=0,
            relief=tk.FLAT
        )
        self.canvas.pack(pady=20)
        self.canvas.bind("<Button-1>", self._on_canvas_click)
        self.canvas.bind("<Motion>", self._on_canvas_motion)
        self.canvas.bind("<Leave>", self._on_canvas_leave)
        
        # 结果显示区域
        result_frame = tk.LabelFrame(
            left_frame,
            text="🎉 今天宠幸它！",
            font=("SimHei", 12, "bold"),
            bg="#f0f0f0",
            fg="#667eea",
            padx=15,
            pady=15
        )
        result_frame.pack(fill=tk.X, pady=10)
        
        self.result_label = tk.Label(
            result_frame,
            text="哪位幸运儿捏...",
            font=("SimHei", 18, "bold"),
            bg="white",
            fg="#667eea",
            pady=20,
            relief=tk.FLAT
        )
        self.result_label.pack(fill=tk.X)
        
        # ========== 右侧：控制面板 ==========
        right_frame = tk.Frame(self.root, bg="#f0f0f0", width=300)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=False, padx=20, pady=20)
        right_frame.pack_propagate(False)
        
        # 标题
        title_label = tk.Label(
            right_frame,
            text="📋 控制面板",
            font=("SimHei", 14, "bold"),
            bg="#f0f0f0",
            fg="#667eea"
        )
        title_label.pack(pady=(0, 20))
        
        # 文件选择按钮
        self.select_btn = tk.Button(
            right_frame,
            text="📁 选择 Excel 文件",
            font=("SimHei", 11, "bold"),
            bg="#5dade2",
            fg="white",
            cursor="hand2",
            command=self._on_select_file,
            relief=tk.FLAT,
            padx=15,
            pady=10
        )
        self.select_btn.pack(fill=tk.X, pady=5)
        
        # 刷新按钮
        self.refresh_btn = tk.Button(
            right_frame,
            text="🔄 刷新数据",
            font=("SimHei", 11, "bold"),
            bg="#78C850",
            fg="white",
            cursor="hand2",
            command=self._on_refresh_btn_click,
            relief=tk.FLAT,
            padx=15,
            pady=10,
            state=tk.DISABLED
        )
        self.refresh_btn.pack(fill=tk.X, pady=5)
        
        # 分隔线
        separator = tk.Frame(right_frame, bg="#ddd", height=1)
        separator.pack(fill=tk.X, pady=15)
        
        # 当前文件显示
        file_label = tk.Label(
            right_frame,
            text="📄 当前文件：",
            font=("SimHei", 10, "bold"),
            bg="#f0f0f0",
            fg="#333"
        )
        file_label.pack(anchor=tk.W, pady=(10, 5))
        
        self.file_display = tk.Label(
            right_frame,
            text="未选择",
            font=("SimHei", 9),
            bg="white",
            fg="#999",
            padx=10,
            pady=8,
            relief=tk.FLAT,
            wraplength=260,
            justify=tk.LEFT
        )
        self.file_display.pack(fill=tk.X, pady=(0, 15))
        
        # 分隔线
        separator2 = tk.Frame(right_frame, bg="#ddd", height=1)
        separator2.pack(fill=tk.X, pady=15)
        
        # 名单标题
        list_title = tk.Label(
            right_frame,
            text="📋 当前名单",
            font=("SimHei", 11, "bold"),
            bg="#f0f0f0",
            fg="#667eea"
        )
        list_title.pack(anchor=tk.W, pady=(0, 10))
        
        # 名单显示区域（带滚动条）
        list_frame = tk.Frame(right_frame, bg="white", relief=tk.FLAT)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.list_text = tk.Listbox(
            list_frame,
            font=("SimHei", 9),
            bg="white",
            fg="#333",
            relief=tk.FLAT,
            yscrollcommand=scrollbar.set,
            cursor="arrow"
        )
        self.list_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.list_text.yview)
        
        # 项目计数
        self.count_label = tk.Label(
            right_frame,
            text="共 0 项",
            font=("SimHei", 9, "bold"),
            bg="#f0f0f0",
            fg="#667eea"
        )
        self.count_label.pack(anchor=tk.W)
    
    def _draw_pointer(self):
        """绘制指针"""
        # 清除旧指针
        self.canvas.delete("pointer")
        
        # 绘制三角形指针（在转盘顶部）
        pointer_size = 20
        points = [
            POINTER_X, POINTER_Y,  # 顶点
            POINTER_X - pointer_size, POINTER_Y + pointer_size * 1.5,  # 左下
            POINTER_X + pointer_size, POINTER_Y + pointer_size * 1.5   # 右下
        ]
        
        self.canvas.create_polygon(
            points,
            fill="#FF6B6B",
            outline="#CC0000",
            width=2,
            tags="pointer"
        )

    def _on_canvas_click(self, event):
        """处理画布点击事件，如果点击的是中心圆则开始旋转"""
        dx = event.x - self.wheel.center_x
        dy = event.y - self.wheel.center_y
        center_radius = 30
        if dx * dx + dy * dy <= center_radius * center_radius:
            self.spin_wheel()

    def _on_canvas_motion(self, event):
        """根据鼠标位置切换光标样式"""
        dx = event.x - self.wheel.center_x
        dy = event.y - self.wheel.center_y
        center_radius = 30
        if dx * dx + dy * dy <= center_radius * center_radius and not self.wheel.is_spinning:
            self.canvas.config(cursor="hand2")
        else:
            self.canvas.config(cursor="arrow")

    def _on_canvas_leave(self, event):
        """鼠标离开画布时恢复默认光标"""
        self.canvas.config(cursor="arrow")
    
    def _on_select_file(self):
        """处理文件选择按钮点击"""
        file_path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[
                ("Excel 文件", "*.xlsx *.xls"),
                ("所有文件", "*.*")
            ]
        )
        
        if file_path:
            self.current_file = file_path
            self._load_excel_data()
    
    def _load_excel_data(self):
        """加载 Excel 数据"""
        if not self.current_file:
            messagebox.showerror("错误", "请先选择 Excel 文件")
            return
        
        try:
            items = ExcelReader.read_first_column(self.current_file)
            
            if not items:
                messagebox.showwarning("警告", "Excel 文件中没有找到有效数据！")
                self.wheel.set_items([])
                self.wheel.draw()
                self.list_text.delete(0, tk.END)
                self.count_label.config(text="共 0 项")
                self.refresh_btn.config(state=tk.DISABLED)
                return
            
            # 更新转盘
            self.wheel.set_items(items)
            self.wheel.reset()
            self.canvas.tag_raise("pointer")
            
            # 更新文件显示
            file_name = Path(self.current_file).name
            self.file_display.config(text=file_name, fg="#333")
            
            # 更新名单显示
            self.list_text.delete(0, tk.END)
            for i, item in enumerate(items, 1):
                self.list_text.insert(tk.END, f"{i}. {item}")
            
            # 更新计数
            self.count_label.config(text=f"共 {len(items)} 项")
            
            # 启用按钮
            self.refresh_btn.config(state=tk.NORMAL)
            
            # 清空结果
            self.result_label.config(text="等待ing...")
            
            messagebox.showinfo("成功", f"成功读取 {len(items)} 条数据！")
        
        except Exception as e:
            messagebox.showerror("错误", str(e))
    
    def _on_refresh_btn_click(self):
        """处理刷新按钮点击"""
        self._load_excel_data()
    
    def spin_wheel(self):
        """开始旋转转盘"""
        if self.wheel.is_spinning or not self.wheel.items:
            return

        self.wheel.is_spinning = True
        self.result_label.config(text="转盘转动中...", fg="#667eea")

        self.spin_start_time = time.time()
        self.spin_duration = 4.0
        self.start_rotation = self.wheel.current_rotation % 360
        rotations = random.randint(5, 10) + random.random()
        self.spin_target = self.start_rotation + rotations * 360

        self.animate_spin()

    def animate_spin(self):
        """使用 Tkinter after() 实现转盘动画"""
        elapsed = time.time() - self.spin_start_time
        progress = min(elapsed / self.spin_duration, 1.0)

        ease_progress = 1 - (1 - progress) ** 3
        self.wheel.current_rotation = self.start_rotation + (self.spin_target - self.start_rotation) * ease_progress
        self.wheel.draw()
        self.canvas.tag_raise("pointer")

        if progress < 1.0:
            self.root.after(16, self.animate_spin)
        else:
            self.wheel.is_spinning = False
            self.calculate_result()

    def calculate_result(self):
        """根据顶部指针位置计算中奖结果并显示"""
        if not self.wheel.items:
            return

        rotation = self.wheel.current_rotation % 360
        slice_angle = 360 / len(self.wheel.items)
        winner_index = int((360 - rotation) / slice_angle) % len(self.wheel.items)
        winner = self.wheel.items[winner_index]
        self.wheel.winner = winner

        self.result_label.config(text=f"今日份阅读指南：{winner}", fg="#667eea")


# ==================== 程序入口 ====================

if __name__ == "__main__":
    root = tk.Tk()
    app = WheelApp(root)
    root.mainloop()
